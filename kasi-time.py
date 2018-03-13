import traceback
import __future__

from openpyxl.utils import get_column_letter

from pyscraper.selenium_utils import get_headed_driver, wait_for_xpath, get_headless_driver
from openpyxl import Workbook


# def date_sequence(date : str, days, backwards=True):
#     months = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
#     year = int(date[0:4])
#     month = int(date[4:6])
#     day = int(date[6:8])
#
#     while days > 0:
#         day -= 1
#         if day == 0:
#             month -= 1
#             if month == 0:
#                 month = 12
#                 year -= 1
#             day = months[month]
#
#         out_month = str(month) if month > 9 else ('0' + str(month))
#         out_day = str(day) if day > 9 else ('0' + str(day))
#         yield (str(year) + out_month + out_day)
#         days -= 1


driver = get_headless_driver(no_sandbox=True)
# for date in date_sequence('20170710', 10):
#     driver.get('http://www.kasi-time.com/day.php?date=' + date)

errors = 0
wb = Workbook()
ws = wb.active
row = 1

ws[get_column_letter(1) + str(row)] = 'Title'
ws[get_column_letter(2) + str(row)] = 'Singer'
ws[get_column_letter(3) + str(row)] = 'Lyrics'
ws[get_column_letter(4) + str(row)] = 'Lyrics Author'
ws[get_column_letter(5) + str(row)] = 'Composition'
ws[get_column_letter(6) + str(row)] = 'Arrangement'
ws[get_column_letter(7) + str(row)] = 'Category'

row += 1

try:
    for i in reversed(range(82179)):
        try:
            driver.get('http://www.kasi-time.com/item-{}.html'.format(str(i)))
            TITLE =         driver.find_elements_by_xpath('/html/body/div/div[2]/div/div/div[1]/div[2]/div/h1')[0].text.encode('utf-8')
            LYRICS =        driver.find_elements_by_xpath('//*[@id="lyrics"]')[0].text.encode('utf-8')
            singer =        driver.find_elements_by_xpath('/html/body/div/div[2]/div/div/div[1]/div[2]/div/div[1]/table/tbody/tr[1]/td/a')[0].text.encode('utf-8')
            lyrics =        driver.find_elements_by_xpath('/html/body/div/div[2]/div/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td/a')[0].text.encode('utf-8')
            composition =   driver.find_elements_by_xpath('/html/body/div/div[2]/div/div/div[1]/div[2]/div/div[1]/table/tbody/tr[3]/td/a')[0].text.encode('utf-8')
            try: arrangement =   driver.find_elements_by_xpath('/html/body/div/div[2]/div/div/div[1]/div[2]/div/div[1]/table/tbody/tr[4]/td/a')[0].text.encode('utf-8')
            except: traceback.print_exc(); arrangement = ''
            try: category =      driver.find_elements_by_xpath('/html/body/div/div[2]/div/div/div[1]/div[2]/div/div[2]/table/tbody/tr[1]/td/a')[0].text.encode('utf-8')
            except: traceback.print_exc(); category = ''
            print(TITLE)

            ws[get_column_letter(1) + str(row)] = TITLE
            ws[get_column_letter(2) + str(row)] = singer
            ws[get_column_letter(3) + str(row)] = LYRICS
            ws[get_column_letter(4) + str(row)] = lyrics
            ws[get_column_letter(5) + str(row)] = composition
            ws[get_column_letter(6) + str(row)] = arrangement
            ws[get_column_letter(7) + str(row)] = category
            row += 1
        except:
            traceback.print_exc()
            errors += 1

        print(i, errors)
finally:
    wb.save('kasi' + '.xlsx')

