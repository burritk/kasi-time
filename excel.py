from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sqlite3

conn = sqlite3.connect('kasi.db')
c = conn.cursor()

wb  = Workbook()
ws = wb.active
for index, row in enumerate(c.execute('select * from lyrics')):
  TITLE, LYRICS, singer, lyrics, composition, arrangement, category = row
  print TITLE
  ws[get_column_letter(1) + str(index)] = TITLE
  ws[get_column_letter(2) + str(index)] = LYRICS
  ws[get_column_letter(3) + str(index)] = singer
  ws[get_column_letter(4) + str(index)] = lyrics
  ws[get_column_letter(5) + str(index)] = composition
  ws[get_column_letter(6) + str(index)] = arrangement
  ws[get_column_letter(7) + str(index)] = category
  
wb.save('test2.xlsx')
